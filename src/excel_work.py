import openpyxl, os, random, re, docx, PIL, docx2txt
# from openpyxl.utils import get_column_letter
import find_rename as fr
import win32com.client as wc

def get_column_values(file_path, column_letters):
    # Load the Excel workbook
    workbook = openpyxl.load_workbook(file_path)
    result_list = []

    # Select the active sheet
    sheet = workbook.active

    # Convert column letters to column indices
    column_indices = [openpyxl.utils.column_index_from_string(col) for col in column_letters]
    # print(column_indices)

    # Iterate over the rows and draw column values
    for row in sheet.iter_rows(values_only=True):
        column_values = [row[col_idx - 1] for col_idx in column_indices]
        result_list.append(column_values)

    # Close the workbook
    workbook.close()

    return result_list

def create_excel(file_path, column_letters):
    input_list = get_column_values(file_path, column_letters)

    # Create a new workbook
    workbook = openpyxl.Workbook()
    output_path = ''
    if os.path.isdir(file_path):
        dir_path = file_path
        output_path = os.path.join(dir_path, ('output' + str(random.randint(1,100)) + '.xlsx'))
    else:
        dir_path = os.path.split(file_path)[0]
        output_path = os.path.join(dir_path, os.path.splitext(os.path.split(file_path)[1])[0] + '_out' + '.xlsx')

    # Get the active sheet
    sheet = workbook.active

    # Iterate over the results list and populate the cells
    for row_idx, row in enumerate(input_list, start=1):
        for col_idx, value in enumerate(row, start=1):
            sheet.cell(row=row_idx, column=col_idx, value=value)

    # Save the workbook to a file
    workbook.save(output_path)

def dir_excels(dir_path, column_letters):
    contents = os.listdir(dir_path)
    for content in contents:
        content_path = os.path.join(dir_path, content)
        if os.path.isfile(content_path):
            create_excel(content_path, column_letters)

def read_image(file_path):
    file_extension = os.path.splitext(file_path)[1]
    dir_path, file_name = os.path.split(file_path)
    n = 1
    if file_extension == '.doc' or file_extension == '.docx':
        # Load the Word document
        doc = docx.Document(file_path)

        ''''
        for k, inline_shape in enumerate(doc.inline_shapes):
            print(type(inline_shape), k)
            try:
                print('test1')
                if isinstance(inline_shape, docx.shape.InlineShape):
                    print('test2')
                    # Get the image from the inline shape object.
                    # image = inline_shape.image
                    image = inline_shape.src
                    print('test3')
                    # Do something with the image.
                    # For example, you can save the image to a file.
                    image.save(f'image{k}.png')

                image = inline_shape.picture

                image_data = image.image_data
                image_filename = image_data.filename
                image_data.save(f"extracted_images/{image_filename}")
            except AttributeError:
                pass

        text = docx2txt.process(file_path)
        text1 = docx2txt.process(file_path, os.path.split(file_path)[0])  # this works
        print(text.splitlines())
        print(text1)
        '''

        word_app = wc.Dispatch("Word.Application")
        doc = word_app.Documents.Open(file_path)
        k = 1
        for shape in doc.Shapes:
            print('test0')
            if shape.Type == 13:
                print('test1')
                image_path = os.path.join(os.path.split(file_path)[0], f"image{k}.png")
                shape.Select()
                print('test1')
                word_app.Selection.Copy()
                img = word_app.Documents.Add()
                img.ActiveWindow.Selection.Paste()
                img.SaveAs(image_path)
                img.Close()
        doc.Close()
        word_app.Quit()


    elif file_extension == '.xls' or file_extension == '.xlsx':
        pass


if __name__ == '__main__':
    # dir_path = input('\nEnter the working directory path: ')
    # file_name = input('Enter the full file name: ')
    # file_path = os.path.join(dir_path, file_name)
    # column_letters_str = input('Enter column letters with comma(,) separator: ')
    # column_letters = re.findall(r'[a-zA-Z]', column_letters_str)
    # column_letters = [char for char in column_letters_str.split(',') if char.isalpha()]

    # res = get_column_values(file_path, column_letters)
    # fr.ask_for_print(res)

    # create_excel(file_path, column_letters)
    # dir_excels(dir_path, column_letters)
    file_path = r'D:\aihub data2\006.한-영 및 한-중 음성발화 데이터\02.데이터\04.데이터\03.데이터\family.docx'
    read_image(file_path)